from flask import Flask, request, jsonify, session, send_from_directory
from flask_pymongo import PyMongo
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_cors import CORS
import os
from dotenv import load_dotenv
import cv2
import PIL.Image
import textwrap
import google.generativeai as genai
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from pathlib import Path
import shutil
import threading
from werkzeug.utils import secure_filename
import os
import PIL.Image
import json
import time
from typing import List
from datetime import datetime


load_dotenv()


GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY')
genai.configure(api_key=GOOGLE_API_KEY)
vision_model = genai.GenerativeModel(model_name="gemini-1.5-pro")
text_model = genai.GenerativeModel('gemini-1.5-flash')
roll_model = genai.GenerativeModel('gemini-1.5-flash')


model = YOLO("model/rectangle_yolo_model.pt")
model_path = YOLO("model/rectangle_yolo_model.pt")


app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "http://localhost:5173"}}, supports_credentials=True)


app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-here')
app.config['MONGO_URI'] = os.getenv('MONGO_URI')


UPLOAD_FOLDER = 'static/uploads'
GENERATED_FOLDER = 'static/generated'
IMAGE_FOLDER = os.path.join(os.getcwd(), 'confirm_images')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['GENERATED_FOLDER'] = GENERATED_FOLDER
app.config['UPLOAD_FOLDER_OMR'] = 'uploads'
app.config['UPLOAD_FOLDER_TEXT'] = 'uploades'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 


os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(GENERATED_FOLDER, exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER_OMR'], exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER_TEXT'], exist_ok=True)
os.makedirs(IMAGE_FOLDER, exist_ok=True)


mongo = PyMongo(app)

processing_status = {}

def to_markdown(text):
    text = text.replace('•', '  *')
    return textwrap.indent(text, '> ', predicate=lambda _: True)

def get_label(image_path, model_path):
    folder_path = 'predict'

  
    if os.path.exists(folder_path):
        try:
            shutil.rmtree(folder_path)
            print(f"Folder '{folder_path}' and all its contents have been deleted.")
        except OSError as e:
            print(f"Error: {e.strerror}")
    else:
        print(f"Folder '{folder_path}' does not exist.")

    output_folder = 'predict'

   
    results = model.predict(image_path, save=True, save_txt=True, project=output_folder, name='results', exist_ok=True)

   
    saved_image_folder = Path(f"{output_folder}/results")
    saved_label_folder = saved_image_folder / 'labels'

    
    saved_labels = list(saved_label_folder.glob('*.txt'))


    if saved_labels:
        label_file_path = saved_labels[0]  
        
        with open(label_file_path, 'r') as file:
            label_data = file.read()

   
        print("Label Data:", label_data)
    else:
        print("No label files found.")
    return label_data


def crop_left_strip(image):
    crop = 30
    height, width = image.shape[:2]
    cropped_image = image[:, crop:]
    return cropped_image


def detect_filled_bubbles(roi):
  
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)


    blurred = cv2.GaussianBlur(gray, (3, 3), 0)


    _, binary = cv2.threshold(blurred, 90, 255, cv2.THRESH_BINARY_INV)


    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    detected_answers = []

   
    min_area = 40  
    max_area = 500
    fill_threshold = 0.3

   
    any_bubble_detected = False

    for contour in contours:
        area = cv2.contourArea(contour)
        if min_area < area < max_area:
            x, y, w, h = cv2.boundingRect(contour)
            bubble_roi = binary[y:y + h, x:x + w]
            filled_area = cv2.countNonZero(bubble_roi)
            if filled_area / (w * h) > fill_threshold:
                
                cx, cy = x + w // 2, y + h // 2
                detected_answers.append((cx, cy))
             
                cv2.circle(roi, (x + w // 2, y + h // 2), 5, (0, 255, 0), 2)
                any_bubble_detected = True

 
    if not any_bubble_detected:
        detected_answers.append((0, 0)) 

    return detected_answers



def convert_to_2d_list(data_str):

    lines = data_str.strip().split('\n')

    data_list = [list(map(float, line.split())) for line in lines]
    return data_list



def final_answers(image_path, data_str):

    save_folder = "confirm_images"
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    image = cv2.imread(image_path)


    height, width, _ = image.shape


    labels = convert_to_2d_list(data_str)

  
    boxes = sorted([label for label in labels if label[0] == 0], key=lambda x: x[1])

    target_width = 150
    target_height = 740

  
    option_ranges = {
        'A': (1, 23),
        'B': (25, 45),
        'C': (48, 67),
        'D': (72, 92),
        'E': (93, 115)
    }


    detected_options = []
    for idx, label in enumerate(boxes):
  
        class_id, center_x, center_y, w, h = label

      
        x_center = int(center_x * width)
        y_center = int(center_y * height)
        box_width = int(w * width)
        box_height = int(h * height)

    
        x1 = int(x_center - box_width / 2)
        y1 = int(y_center - box_height / 2)
        x2 = int(x_center + box_width / 2)
        y2 = int(y_center + box_height / 2)


        roi = image[y1:y2, x1:x2].copy()
        roi_resize = cv2.resize(roi, (target_width, target_height))
        roi = crop_left_strip(roi_resize)


        section_height = roi.shape[0] / 50.0

   
        for j in range(1, 50):
            y_line = int(j * section_height)
            cv2.line(roi, (0, y_line), (roi.shape[1], y_line), (0, 255, 0), 1)

        
        cv2.line(roi, (0, roi.shape[0] - 1), (roi.shape[1], roi.shape[0] - 1), (0, 255, 0), 1)

        
        for j in range(50):
            y_start = int(j * section_height)
            y_end = int((j + 1) * section_height)

            
            if j == 49:
                y_end = roi.shape[0]

            section = roi[y_start:y_end, :]

            
            detected_answers = detect_filled_bubbles(section)

          
            for cx, cy in detected_answers:
                for option, (min_x, max_x) in option_ranges.items():
                    if min_x <= cx < max_x:
                        detected_options.append(option)
                        break
                else:
                    
                    detected_options.append('0')

        
        save_path = os.path.join(save_folder, f"rect_{idx + 1}.jpg")
        cv2.imwrite(save_path, roi)

    
    print(len(detected_options))

    return detected_options


def find_rollnumber(image_path: str, label_data: str) -> str:
    response = "Roll No."
   
    image = cv2.imread(image_path)

    
    height, width, _ = image.shape

  
    def convert_to_2d_list(data_str: str) -> List[List[float]]:
        lines = data_str.strip().split('\n')
        data_list = [list(map(float, line.split())) for line in lines]
        return data_list

    labels = convert_to_2d_list(label_data)

    # Filter the boxes with class 1 and save ROI
    for label in labels:
        class_id, center_x, center_y, w, h = label
        if class_id == 1:
            x_center = int(center_x * width)
            y_center = int(center_y * height)
            box_width = int(w * width)
            box_height = int(h * height)

            x1 = int(x_center - box_width / 2)
            y1 = int(y_center - box_height / 2)
            x2 = int(x_center + box_width / 2)
            y2 = int(y_center + box_width / 2)

            # Save the ROI as "roll.jpg"
            roi = image[y1:y2, x1:x2]
            cv2.imwrite("roll1.jpg", roi)
            prompt = """
            Task: Extract the 9-digit roll number from the provided image. The digits are enclosed within 9 separate boxes, so the following constraints must be strictly followed to ensure maximum accuracy.

            Accuracy Requirement:
            You must extract all 9 digits from the image with 100% consistency. The output must not change if the image is processed multiple times.
            If possible, store or cache the result of the extraction so that the output remains identical across multiple requests for the same image.
            Digit Recognition:
            Each of the 9 boxes contains exactly one digit. Ensure that you check every box to extract exactly 9 digits.
            Carefully recognize each digit inside its respective box without confusing the boundaries of the box with the digit.
            Special care must be taken to distinguish between commonly confused digits such as 0 and 6. Use boundary detection techniques to ensure digits are not misclassified due to poor alignment with the box.
            No digit should be skipped or misclassified. Ensure that the digit inside every box is read and extracted correctly.
            Additional Techniques:
            Use advanced methods like image pre-processing, boundary isolation, or digit enhancement techniques to ensure that the extraction process is accurate and consistent.
            Implement digit-box separation algorithms to help prevent misinterpretation of digits due to overlapping boundaries or poor contrast.
            Critical Review:
            After extraction, perform an internal validation pass to ensure that exactly 9 digits are extracted, and that they are correct based on the image input.
            Key Considerations:
            Consistency: The same image must always yield the same result, stored if necessary to avoid variability.
            Digit Separation: Ensure proper distinction between digits and box boundaries, especially when dealing with ambiguous digits like "0" and "6".
            No Missing Digits: Ensure that all 9 digits are extracted from their corresponding boxes.
            Explanation:
            Emphasis on 9 boxes and 9 digits ensures that each digit is accounted for and no box is skipped.
            Consistency and caching instructions help mitigate the variability of outputs for repeated image processing.
            Boundary detection and digit enhancement are reiterated to avoid common misclassification errors, particularly between "0" and "6".
            Internal validation guarantees that after extraction, no digit is missed or incorrectly processed, ensuring all 9 boxes are captured properly.
            **Do not write anything from your side just provide the detected roll number as output**
            """
            roll_image = PIL.Image.open('roll1.jpg')
            response = roll_model.generate_content([prompt, roll_image])
            response = response.text
            print(response)

    return response
def chq(detected_list, questions):
    words = list(questions)
    words = [item.lower() for item in words]
    detected_list = [item.lower() for item in detected_list]

    wrong_ans = []
    
    total_marks = 0
    for i in range(len(words)):
        if words[i] == detected_list[i]:
            total_marks += 1
        else:
            wrong_ans.append(i + 1)
    total_marks = str(total_marks) + " / " + str(len(words))

    wrong_answer = "<br>".join(f"Question {i}" for i in wrong_ans)
    return total_marks, wrong_answer

def feed_record(roll_number, marks_obtained):
    # File path
    excel_file = "student_marks.xlsx"

    # Create or load the Excel sheet
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        # Add headers
        sheet["A1"] = "Roll Number"
        sheet["B1"] = "Marks Obtained"

    # Function to check if roll number already exists
    def roll_number_exists(roll_number):
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == roll_number:
                return True
        return False

    # Function to add entry
    def add_entry(roll_number, marks_obtained):
        if not roll_number_exists(roll_number):
            # Find the next available row
            next_row = sheet.max_row + 1
            # Store data
            sheet[f"A{next_row}"] = roll_number
            sheet[f"B{next_row}"] = marks_obtained
            # Save the workbook
            workbook.save(excel_file)
            print(f"Entry added: Roll Number: {roll_number}, Marks Obtained: {marks_obtained}")
        else:
            print(f"Duplicate entry avoided for Roll Number: {roll_number}.")
    add_entry(roll_number, marks_obtained)

roll_num=None
total_marks=None
wrong_answers=None
# API Routes
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')

    if not all([name, email, password]):
        return jsonify({"success": False, "message": "All fields are required"}), 400

    if mongo.db.users.find_one({"email": email}):
        print("user already exist")
        return jsonify({"success": False, "message": "Email already registered"}), 400

    hashed_password = generate_password_hash(password)
    mongo.db.users.insert_one({
        "name": name,
        "email": email,
        "password": hashed_password
    })

    print("user registered successfully")

    return jsonify({
        "success": True,
        "message": "Registration successful",
        "user": {
            "name": name,
            "email": email
        }
    })

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    if not all([email, password]):
        return jsonify({"success": False, "message": "Email and password are required"}), 400

    user = mongo.db.users.find_one({"email": email})
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404

    if not check_password_hash(user['password'], password):
        return jsonify({"success": False, "message": "Invalid credentials"}), 401

    session['email'] = email
    return jsonify({
        "success": True,
        "message": "Login successful",
        "user": {
            "name": user['name'],
            "email": user['email']
        }
    })

@app.route('/api/logout', methods=['POST'])
def logout():
    session.pop('email', None)
    return jsonify({"success": True, "message": "Logged out successfully"})

@app.route('/api/me')
def check_auth():
    if 'email' not in session:
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    user = mongo.db.users.find_one({"email": session['email']})
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404

    return jsonify({
        "success": True,
        "user": {
            "name": user['name'],
            "email": user['email']
        }
    })

@app.route('/api/upload-omr', methods=['POST'])
def upload_omr():
    if 'email' not in session:
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400

    if not file:
        return jsonify({"success": False, "message": "Invalid file"}), 400

    answer_key = request.form.get("question")
    if not answer_key:
        return jsonify({"success": False, "message": "Answer key is required"}), 400

    try:
        # Save the uploaded file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER_OMR'], filename)
        file.save(file_path)

        # Process the OMR sheet
        labels = get_label(file_path, model_path)
        roll_num = find_rollnumber(file_path, labels)
        detected_answers = final_answers(file_path, labels)
        
        # Clean the answer key (remove spaces, make lowercase)
        answer_key_clean = [x.strip().lower() for x in answer_key.split(',')]
        
        # Calculate score
        score = 0
        wrong_answers = []
        
        for i in range(min(len(detected_answers), len(answer_key_clean))):
            if detected_answers[i].lower() == answer_key_clean[i]:
                score += 1
            else:
                wrong_answers.append(i+1)  # 1-based question numbers

        total_questions = len(answer_key_clean)
        score_str = f"{score}/{total_questions}"

        # Save the result
        feed_record(roll_num, score_str)

        return jsonify({
            "success": True,
            "detected_answers": detected_answers,
            "roll_number": roll_num,
            "total_marks": score_str,
            "wrong_answers": len(wrong_answers),
            "wrong_answer_list": wrong_answers
        })

    except Exception as e:
        app.logger.error(f"Error processing OMR: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Error processing OMR: {str(e)}"
        }), 500

@app.route('/api/evaluate-text', methods=['POST'])
def evaluate_text():
    print("Evaluating the answer")
    if 'email' not in session:
        return jsonify({"success": False, "message": "Not authenticated"}), 401

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400

    question = request.form.get('question')
    if not question:
        return jsonify({"success": False, "message": "Question is required"}), 400

    # Generate a unique processing ID
    processing_id = f"eval_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    
    # Initialize processing status with a lock
    with threading.Lock():
        processing_status[processing_id] = {
            'status': 'processing',
            'start_time': time.time(),
            'question': question,
            'progress': 0  # 0-100%
        }

    try:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER_TEXT'], filename)
        file.save(file_path)

        def process_evaluation():
            try:
                # Update progress
                with threading.Lock():
                    processing_status[processing_id]['progress'] = 10
                    processing_status[processing_id]['stage'] = 'extracting_text'

                # Text extraction
                image = PIL.Image.open(file_path)
                response = vision_model.generate_content(["Extract the text exactly as written in this image:", image])
                response.resolve()
                extracted_text = response.text

                with threading.Lock():
                    processing_status[processing_id]['progress'] = 40
                    processing_status[processing_id]['stage'] = 'evaluating'

                # Strict evaluation prompt
                evaluation_prompt = f"""Act as a strict examiner evaluating a student's answer. Follow these guidelines:

**Evaluation Rules:**
1. Start with 0 points
2. Only award points for explicitly correct information
3. Give zero if the answer is completely wrong or irrelevant to the question
4. For partially correct answers:
   - 1-3 points: Minimal relevant content
   - 4-6 points: Some correct concepts with errors
   - 7-9 points: Mostly correct with minor gaps
   - 10 points: Perfect answer

**Question:** {question}
**Student's Answer:** {extracted_text}

**Scoring Criteria:**
- Each fully correct key concept: +1-2 points
- Partial understanding: +0.5 point (only if somewhat relevant)
- Scientific accuracy: +1 point
- Complete explanation: +1 point

**Required JSON Output:**
{{
    "score": 0.0,  // 0 for wrong, 1-10 for correct
    "strengths": [],  // Only if points > 0
    "improvements": [],  // Specific corrections needed
    "feedback": ""  // Clear explanation of grading
}}

**Important Notes:**
- If the answer doesn't address the question at all, give 0
- Don't give points for effort - only for correct content
- Be strict but fair in evaluation

Now evaluate this answer strictly:"""
                
                evaluation_response = text_model.generate_content(evaluation_prompt)
                response_text = evaluation_response.text

                # Robust JSON extraction
                try:
                    json_start = response_text.find('{')
                    json_end = response_text.rfind('}') + 1
                    if json_start == -1 or json_end == 0:
                        raise ValueError("No JSON found in response")
                    
                    evaluation = json.loads(response_text[json_start:json_end])
                    
                    # Validate and enforce strict scoring
                    required_keys = {'score', 'strengths', 'improvements', 'feedback'}
                    if not all(key in evaluation for key in required_keys):
                        raise ValueError("Missing required evaluation keys")

                    # Force zero score if completely irrelevant
                    question_keywords = set(question.lower().split())
                    answer_keywords = set(extracted_text.lower().split())
                    if not question_keywords.intersection(answer_keywords):
                        evaluation = {
                            "score": 0.0,
                            "strengths": [],
                            "improvements": ["Your answer doesn't address the question"],
                            "feedback": "This response is completely irrelevant to the question asked."
                        }
                    
                except Exception as e:
                    print(f"Failed to parse evaluation: {str(e)}")
                    evaluation = {
                        "score": 0.0,
                        "strengths": [],
                        "improvements": ["Technical evaluation failed"],
                        "feedback": "Unable to assess answer - please ensure it addresses the question"
                    }

                with threading.Lock():
                    processing_status[processing_id].update({
                        'status': 'completed',
                        'result': {
                            "success": True,
                            "prediction_text": extracted_text,
                            "evaluation": evaluation
                        },
                        'processing_time': time.time() - processing_status[processing_id]['start_time'],
                        'progress': 100
                    })

            except Exception as e:
                with threading.Lock():
                    processing_status[processing_id].update({
                        'status': 'error',
                        'error': str(e),
                        'processing_time': time.time() - processing_status[processing_id]['start_time'],
                        'progress': 100
                    })

        # Start the processing thread
        threading.Thread(target=process_evaluation, daemon=True).start()

        # Return immediate response with processing details
        return jsonify({
            "success": True,
            "message": "Evaluation started",
            "processing_id": processing_id,
            "status_url": f"/api/check-status/{processing_id}",
            "estimated_time": 30  # seconds
        }), 202

    except Exception as e:
        with threading.Lock():
            processing_status[processing_id] = {
                'status': 'error',
                'error': str(e),
                'processing_time': 0
            }
        return jsonify({
            "success": False,
            "message": "Initial processing failed",
            "error": str(e),
            "processing_id": processing_id
        }), 500

@app.route('/api/check-status/<processing_id>', methods=['GET'])
def check_status(processing_id):
    status = processing_status.get(processing_id, {'status': 'not_found'})
    
    if status['status'] == 'completed':
        return jsonify(status['result'])
    elif status['status'] == 'error':
        return jsonify({
            "success": False,
            "message": "Evaluation failed",
            "error": status.get('error', 'Unknown error')
        }), 500
    elif status['status'] == 'processing':
        return jsonify({
            "success": True,
            "status": "processing",
            "elapsed_time": time.time() - status['start_time']
        }), 202
    else:
        return jsonify({
            "success": False,
            "message": "Processing ID not found"
        }), 404




@app.route('/api/results/<filename>')
def get_result_image(filename):
    return send_from_directory(IMAGE_FOLDER, filename)

if __name__ == "__main__":
    app.run(debug=True)